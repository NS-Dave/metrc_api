"""
IRS Audit Export — METRC seed-to-sale for 140 Industrial Road LLC (IDR TC-1, item #26)
========================================================================================
140 Industrial Road LLC is the group's cultivator (MC281599) AND product
manufacturer (MP281433) — confirmed by Dave 7/1/26. This production pipeline
(metrc_daily_sync.py) already syncs both licenses into Supabase, so this pulls
straight from there instead of re-crawling the METRC API.

For TY2023 and TY2024, writes to C:\\python\\irs_audit_2023\\:
    140_<year>_METRC_Summary.xlsx   — monthly Plants / Harvests / Packages / Transfers
    output_140_metrc/140_<year>_metrc_{plants,harvests,packages,transfers}_detail.csv
    140_METRC_Locations.xlsx        — full facility location roster (supports IDR #28/#29:
                                       grow rooms + mother/clone/veg + extraction/manufacturing
                                       + drying/trim/packaging — all indoor, no outdoor/greenhouse
                                       location names found)

Usage:
    cd C:\\python\\metrc_api
    python build_metrc_140_report.py
"""
import os
import sys
import csv

sys.path.insert(0, os.path.dirname(__file__))
from supabase_config import get_connection_string
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IRS_DIR = r"C:\python\irs_audit_2023"
OUT_CSV_DIR = os.path.join(IRS_DIR, "output_140_metrc")
LICENSES = ("MC281599", "MP281433")
YEARS = (2023, 2024)

NAVY, BLUE, WHITE, GREY = "1F3864", "2E5496", "FFFFFF", "F2F2F2"
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def get_conn():
    return psycopg2.connect(get_connection_string())


def connection_ids(cur):
    cur.execute('SELECT id FROM metrc_connections WHERE "licenseNumber" = ANY(%s)', (list(LICENSES),))
    return [r[0] for r in cur.fetchall()]


def style_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row, i, h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def write_rows(ws, start_row, rows):
    r = start_row
    for row in rows:
        for i, val in enumerate(row, start=1):
            c = ws.cell(r, i, val)
            c.border = BORDER
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
        r += 1
    return r


def dump_csv(path, cur, sql, params):
    cur.execute(sql, params)
    cols = [d.name for d in cur.description]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(cur.fetchall())
    return cur.rowcount if cur.rowcount is not None else 0


def main():
    os.makedirs(OUT_CSV_DIR, exist_ok=True)
    conn = get_conn()
    cur = conn.cursor()
    ids = connection_ids(cur)
    print(f"140 Industrial Road connection ids: {ids}")

    for year in YEARS:
        y0, y1 = f"{year}-01-01", f"{year+1}-01-01"
        print(f"\n=== FY{year} ===")

        # ---- detail CSVs (full seed-to-sale record) ----
        n = dump_csv(
            os.path.join(OUT_CSV_DIR, f"140_{year}_metrc_plants_detail.csv"), cur,
            'SELECT "plantLabel","plantState","growthPhase","strainName","locationName",'
            '"plantedDate","vegetativeDate","floweringDate","harvestedDate","destroyedDate",'
            '"plantBatchName","plantBatchType" FROM metrc_plants '
            'WHERE "metrcConnectionId"::text = ANY(%s) AND "plantedDate" >= %s AND "plantedDate" < %s '
            'ORDER BY "plantedDate"', (ids, y0, y1))
        print(f"  plants detail: {n} rows")

        n = dump_csv(
            os.path.join(OUT_CSV_DIR, f"140_{year}_metrc_harvests_detail.csv"), cur,
            'SELECT "harvestName","harvestType","strainName","dryingLocationName","currentWeight",'
            '"unitOfWeight","isFinished","harvestStartDate","finishedDate","sourcePlantCount",'
            '"totalWetWeight","totalWasteWeight","packageCount" FROM metrc_harvests '
            'WHERE "metrcConnectionId"::text = ANY(%s) AND "harvestStartDate" >= %s AND "harvestStartDate" < %s '
            'ORDER BY "harvestStartDate"', (ids, y0, y1))
        print(f"  harvests detail: {n} rows")

        n = dump_csv(
            os.path.join(OUT_CSV_DIR, f"140_{year}_metrc_packages_detail.csv"), cur,
            'SELECT "packageLabel","packageType","packageState","itemName","itemCategory",'
            '"productCategoryName","quantity","unitOfMeasure","initialQuantity","locationName",'
            '"packagedDate","unitCostPrice","totalCost","sourceHarvestNames","labTestingState" '
            'FROM metrc_packages WHERE "metrcConnectionId"::text = ANY(%s) '
            'AND "packagedDate" >= %s AND "packagedDate" < %s ORDER BY "packagedDate"',
            (ids, y0, y1))
        print(f"  packages detail: {n} rows")

        n = dump_csv(
            os.path.join(OUT_CSV_DIR, f"140_{year}_metrc_transfers_detail.csv"), cur,
            'SELECT "manifestNumber","transferType","shipmentTypeName","shipperFacilityLicenseNumber",'
            '"shipperFacilityName","destinationFacilityLicenseNumber","destinationFacilityName",'
            '"transferState","createdDate","shippedDate","receivedDate","packageCount" '
            'FROM metrc_transfers WHERE "metrcConnectionId"::text = ANY(%s) '
            'AND "receivedDate" >= %s AND "receivedDate" < %s ORDER BY "receivedDate"',
            (ids, y0, y1))
        print(f"  transfers detail: {n} rows")

        # ---- monthly summary workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Plants"
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"140 Industrial Road LLC — METRC Plants by Month & Growth Phase (FY{year})"
        ws["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws["A2"] = "IDR #26 — inventory continually accounted for, purchase (planting) through sale"
        ws["A2"].font = Font(italic=True, color=BLUE)
        style_header(ws, 4, ["Month", "Growth phase", "Plant count"], [12, 20, 14])
        cur.execute(
            'SELECT to_char("plantedDate",\'YYYY-MM\') mo, "growthPhase", count(*) '
            'FROM metrc_plants WHERE "metrcConnectionId"::text = ANY(%s) '
            'AND "plantedDate" >= %s AND "plantedDate" < %s GROUP BY 1,2 ORDER BY 1,2',
            (ids, y0, y1))
        write_rows(ws, 5, cur.fetchall())

        ws2 = wb.create_sheet("Harvests")
        ws2.sheet_view.showGridLines = False
        ws2["A1"] = f"140 Industrial Road LLC — METRC Harvests by Month (FY{year})"
        ws2["A1"].font = Font(bold=True, size=13, color=NAVY)
        style_header(ws2, 3, ["Month", "Harvest count", "Total wet weight", "Total waste weight", "Current weight (post-dry)"], [12, 14, 18, 18, 20])
        cur.execute(
            'SELECT to_char("harvestStartDate",\'YYYY-MM\') mo, count(*), '
            'sum("totalWetWeight"), sum("totalWasteWeight"), sum("currentWeight") '
            'FROM metrc_harvests WHERE "metrcConnectionId"::text = ANY(%s) '
            'AND "harvestStartDate" >= %s AND "harvestStartDate" < %s GROUP BY 1 ORDER BY 1',
            (ids, y0, y1))
        write_rows(ws2, 4, cur.fetchall())

        ws3 = wb.create_sheet("Packages")
        ws3.sheet_view.showGridLines = False
        ws3["A1"] = f"140 Industrial Road LLC — METRC Packages by Month & Product Category (FY{year})"
        ws3["A1"].font = Font(bold=True, size=13, color=NAVY)
        style_header(ws3, 3, ["Month", "Product category", "Package count", "Total quantity", "Total cost"], [12, 26, 14, 16, 16])
        cur.execute(
            'SELECT to_char("packagedDate",\'YYYY-MM\') mo, "productCategoryName", count(*), '
            'sum(quantity), sum("totalCost") FROM metrc_packages '
            'WHERE "metrcConnectionId"::text = ANY(%s) AND "packagedDate" >= %s AND "packagedDate" < %s '
            'GROUP BY 1,2 ORDER BY 1,2', (ids, y0, y1))
        write_rows(ws3, 4, cur.fetchall())

        ws4 = wb.create_sheet("Transfers")
        ws4.sheet_view.showGridLines = False
        ws4["A1"] = f"140 Industrial Road LLC — METRC Outgoing Transfers by Month & Destination (FY{year})"
        ws4["A1"].font = Font(bold=True, size=13, color=NAVY)
        style_header(ws4, 3, ["Month", "Destination facility", "Transfer count", "Total packages"], [12, 34, 14, 16])
        cur.execute(
            'SELECT to_char("receivedDate",\'YYYY-MM\') mo, "destinationFacilityName", count(*), '
            'sum("packageCount") FROM metrc_transfers '
            'WHERE "metrcConnectionId"::text = ANY(%s) AND "receivedDate" >= %s AND "receivedDate" < %s '
            'GROUP BY 1,2 ORDER BY 1,2', (ids, y0, y1))
        write_rows(ws4, 4, cur.fetchall())

        out_path = os.path.join(IRS_DIR, f"140_{year}_METRC_Summary.xlsx")
        wb.save(out_path)
        print(f"  wrote {out_path}")

    # ---- location roster (supports #28 cultivation description + #29 site map) ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "140 Industrial Road LLC — METRC Location / Room Roster"
    ws["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws["A2"] = "Supports IDR #28 (description of cultivation activities) and #29 (facility layout) — draft only, confirm with ops"
    ws["A2"].font = Font(italic=True, color=BLUE)
    style_header(ws, 4, ["Location / room name", "Active"], [46, 10])
    cur.execute(
        'SELECT name, "isActive" FROM metrc_locations WHERE "metrcConnectionId"::text = ANY(%s) ORDER BY name',
        (ids,))
    write_rows(ws, 5, cur.fetchall())
    out_path = os.path.join(IRS_DIR, "140_METRC_Locations.xlsx")
    wb.save(out_path)
    print(f"\nwrote {out_path}")

    cur.close()
    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
